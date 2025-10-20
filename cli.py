#!/usr/bin/env python3
"""
eCourts Scraper – Command-Line Interface (Fixed for real eCourts v6)
"""

from __future__ import annotations
import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path

from scraper import eCourtsScraper
import hierarchy_fetcher as hf

# ====================================================================
# Helpers
# ====================================================================
def save_json(data, filename):
    Path("output").mkdir(exist_ok=True)
    filepath = filename if filename.startswith("output/") else f"output/{filename}"
    with open(filepath, "w", encoding="utf-8") as fp:
        json.dump(data, fp, indent=2, ensure_ascii=False)
    print(f"✓ Saved to: {filepath}")

def save_text(data, filename):
    Path("output").mkdir(exist_ok=True)
    filepath = filename if filename.startswith("output/") else f"output/{filename}"
    with open(filepath, "w", encoding="utf-8") as fp:
        if data.get("data") and isinstance(data["data"], dict):
            if "cases" in data["data"]:
                fp.write(f"CAUSE LIST – {data['data']['date']}\n")
                fp.write(f"Total Cases: {data['data']['total_cases']}\n")
                fp.write("=" * 80 + "\n\n")
                for case in data["data"]["cases"]:
                    fp.write(f"Serial No: {case.get('serial_number')}\n")
                    fp.write(f"Case Number: {case.get('case_number')}\n")
                    fp.write(f"Parties: {case.get('parties')}\n")
                    fp.write(f"Purpose: {case.get('purpose')}\n")
                    fp.write("-" * 80 + "\n")
            else:
                fp.write("CASE DETAILS\n")
                fp.write("=" * 80 + "\n\n")
                for key, value in data["data"].items():
                    if value:
                        fp.write(f"{key.replace('_', ' ').title()}: {value}\n")
    print(f"✓ Saved to: {filepath}")

def export_excel(data, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("⚠ Excel export requires openpyxl (pip install openpyxl)")
        return False

    if not (data.get("data") and "cases" in data["data"]):
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Cause List"

    headers = ["Sr.No", "Case Number", "Parties", "Purpose", "Court"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for case in data["data"]["cases"]:
        ws.append([
            case.get("serial_number", ""),
            case.get("case_number", ""),
            case.get("parties", ""),
            case.get("purpose", ""),
            case.get("court_name", ""),
        ])

    for column in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)

    wb.save(filename)
    print(f"✓ Excel file saved: {filename}")
    return True

def display_case(result):
    if not result["success"]:
        print(f"\n❌ {result['message']}\n")
        return

    c = result["data"]
    print("\n" + "=" * 70)
    print("CASE DETAILS")
    print("=" * 70)
    print(f"Case Number: {c.get('case_number', 'N/A')}")
    print(f"Case Type: {c.get('case_type', 'N/A')}")
    print(f"Filing Date: {c.get('filing_date', 'N/A')}")
    print(f"Court: {c.get('court_name', 'N/A')}")
    print(f"Judge: {c.get('judge_name', 'N/A')}")
    print(f"Petitioner: {c.get('petitioner', 'N/A')}")
    print(f"Respondent: {c.get('respondent', 'N/A')}")
    print(f"Next Hearing: {c.get('next_hearing', 'N/A')}")
    print(f"Status: {c.get('status', 'N/A')}")
    print("\nLISTING STATUS:")
    if c.get("is_listed_today"):
        print(f"✓ Listed TODAY – Serial No: {c.get('serial_number', 'N/A')}")
    elif c.get("is_listed_tomorrow"):
        print(f"✓ Listed TOMORROW – Serial No: {c.get('serial_number', 'N/A')}")
    else:
        print("✗ Not listed today or tomorrow")
    print("=" * 70 + "\n")

def display_cause_list(result):
    if not result["success"]:
        print(f"\n❌ {result['message']}\n")
        return

    data = result["data"]
    print("\n" + "=" * 100)
    print(f"CAUSE LIST – {data['date']} ({data['total_cases']} cases)")
    print("=" * 100)
    print(f"{'Sr.No':<8} {'Case Number':<20} {'Parties':<40} {'Purpose':<20}")
    print("-" * 100)
    for case in data["cases"]:
        print(
            f"{case.get('serial_number', ''):<8} "
            f"{case.get('case_number', '')[:19]:<20} "
            f"{case.get('parties', '')[:39]:<40} "
            f"{case.get('purpose', '')[:19]:<20}"
        )
    print("=" * 100 + "\n")

def display_statistics(result):
    if not result["success"]:
        return

    data = result["data"]
    cases = data.get("cases", [])

    if not cases:
        print("No cases to analyze")
        return

    # Calculate stats
    stats = {
        "total_cases": len(cases),
        "by_purpose": {},
        "by_court": {},
    }

    for case in cases:
        purpose = case.get("purpose", "Unknown")
        court = case.get("court_name", "Unknown")

        stats["by_purpose"][purpose] = stats["by_purpose"].get(purpose, 0) + 1
        stats["by_court"][court] = stats["by_court"].get(court, 0) + 1

    stats["unique_purposes"] = len(stats["by_purpose"])
    stats["unique_courts"] = len(stats["by_court"])

    # Display
    print("\n" + "=" * 70)
    print("STATISTICS")
    print("=" * 70)
    print(f"Total Cases: {stats['total_cases']}")
    print(f"Unique Purposes: {stats['unique_purposes']}")
    print(f"Unique Courts: {stats['unique_courts']}")

    print("\nTop Purposes:")
    sorted_purposes = sorted(stats["by_purpose"].items(), key=lambda x: x[1], reverse=True)[:10]
    for purpose, count in sorted_purposes:
        pct = (count / stats["total_cases"]) * 100
        bar = "█" * int(pct / 2)
        print(f" {purpose[:35]:<35} {count:>3} ({pct:5.1f}%) {bar}")

    if stats["by_court"]:
        print("\nCases by Court:")
        sorted_courts = sorted(stats["by_court"].items(), key=lambda x: x[1], reverse=True)[:10]
        for court, count in sorted_courts:
            pct = (count / stats["total_cases"]) * 100
            print(f" {court[:40]:<40} {count:>3} ({pct:5.1f}%)")

    print("=" * 70 + "\n")

def get_captcha(scraper):
    """Helper to get captcha from user."""
    print("\nFetching captcha...")
    captcha_url = f"{scraper.BASE_URL}/?p=casestatus/getCaptcha"
    code = scraper._get_captcha_code(captcha_url)
    return code

def interactive_hierarchy():
    """Interactive mode to select state/district/complex/court."""
    print("\n" + "=" * 70)
    print("INTERACTIVE COURT SELECTION")
    print("=" * 70)

    # Get states
    print("\nFetching states...")
    states = hf.states()
    if not states:
        print("Could not fetch states")
        return None

    print("\nAvailable States:")
    for i, s in enumerate(states[:20], 1):  # Show first 20
        print(f" {i}. {s['name']} (code: {s['code']})")

    choice = input("\nEnter state number (or 'q' to quit): ").strip()
    if choice.lower() == 'q':
        return None

    try:
        state = states[int(choice) - 1]
    except (ValueError, IndexError):
        print("Invalid choice")
        return None

    # Get districts
    print(f"\nFetching districts for {state['name']}...")
    districts = hf.districts(state['code'])
    if not districts:
        print("Could not fetch districts")
        return None

    print("\nAvailable Districts:")
    for i, d in enumerate(districts[:20], 1):
        print(f" {i}. {d['name']} (code: {d['code']})")

    choice = input("\nEnter district number: ").strip()
    try:
        district = districts[int(choice) - 1]
    except (ValueError, IndexError):
        print("Invalid choice")
        return None

    # Get complexes
    print(f"\nFetching court complexes for {district['name']}...")
    complexes = hf.complexes(state['code'], district['code'])
    if not complexes:
        print("Could not fetch complexes")
        return None

    print("\nAvailable Court Complexes:")
    for i, c in enumerate(complexes[:20], 1):
        print(f" {i}. {c['name']} (code: {c['code']})")

    choice = input("\nEnter complex number: ").strip()
    try:
        complex_obj = complexes[int(choice) - 1]
    except (ValueError, IndexError):
        print("Invalid choice")
        return None

    # Get courts
    print(f"\nFetching courts in {complex_obj['name']}...")
    courts = hf.courts(state['code'], district['code'], complex_obj['code'])
    if not courts:
        print("Could not fetch courts")
        return None

    print("\nAvailable Courts:")
    for i, ct in enumerate(courts[:20], 1):
        print(f" {i}. {ct['name']} (code: {ct['code']})")

    choice = input("\nEnter court number: ").strip()
    try:
        court = courts[int(choice) - 1]
    except (ValueError, IndexError):
        print("Invalid choice")
        return None

    return {
        "state_code": state['code'],
        "dist_code": district['code'],
        "court_complex_code": complex_obj['code'],
        "court_code": court['code'],
        "state_name": state['name'],
        "district_name": district['name'],
        "complex_name": complex_obj['name'],
        "court_name": court['name']
    }

# ====================================================================
# Main
# ====================================================================
def main():
    p = argparse.ArgumentParser(
        description="eCourts Scraper – fetch case info / cause lists",
        epilog='Example: python cli.py --interactive --today --kind civ',
    )

    # Search options
    p.add_argument("--cnr", help="Search by CNR number")
    p.add_argument("--case-type", help="Case type (for case number search)")
    p.add_argument("--case-number", help="Case number")
    p.add_argument("--year", help="Case year")

    # Location codes
    p.add_argument("--state-code", help="State code")
    p.add_argument("--dist-code", help="District code")
    p.add_argument("--complex-code", help="Court complex code")
    p.add_argument("--court-code", help="Court/establishment code")
    p.add_argument("--est-code", default="null", help='Establishment code (default "null")')

    # Cause list options
    p.add_argument("--today", action="store_true", help="Get today's cause list")
    p.add_argument("--tomorrow", action="store_true", help="Get tomorrow's cause list")
    p.add_argument("--causelist", action="store_true", help="Get cause list for date")
    p.add_argument("--date", help="Date (DD-MM-YYYY)")
    p.add_argument("--kind", choices=["civ", "crim"], default="civ", help="Cause list type (civil/criminal)")

    # Utilities
    p.add_argument("--interactive", action="store_true", help="Interactive mode to select court")
    p.add_argument("--captcha", help="Captcha code (if already solved)")
    p.add_argument("--no-ocr", action="store_true", help="Disable OCR, use manual captcha entry")
    p.add_argument("--stats", action="store_true", help="Show statistics")
    p.add_argument("--download-pdf", action="store_true", help="Download PDF if available")

    # Output options
    p.add_argument("--output-json", help="Save as JSON file")
    p.add_argument("--output-text", help="Save as text file")
    p.add_argument("--excel", help="Save as Excel file")

    args = p.parse_args()
    scraper = eCourtsScraper(use_ocr=not args.no_ocr)
    result = None

    # Get location if interactive mode
    location = None
    if args.interactive:
        location = interactive_hierarchy()
        if not location:
            print("Interactive selection cancelled")
            return

        print("\nSelected:")
        print(f"  State: {location['state_name']}")
        print(f"  District: {location['district_name']}")
        print(f"  Complex: {location['complex_name']}")
        print(f"  Court: {location['court_name']}")

    # Override with location if available
    if location:
        args.state_code = location['state_code']
        args.dist_code = location['dist_code']
        args.complex_code = location['court_complex_code']
        args.court_code = location['court_code']

    # ================== single-search paths ==================
    if args.cnr:
        captcha = args.captcha or get_captcha(scraper)
        if not captcha:
            print("Captcha required")
            return

        result = scraper.search_by_cnr(args.cnr, captcha)
        display_case(result)

    elif args.case_type and args.case_number and args.year:
        if not all([args.state_code, args.dist_code, args.complex_code]):
            print("Need --state-code, --dist-code, and --complex-code")
            print(" Or use --interactive mode")
            return

        captcha = args.captcha or get_captcha(scraper)
        if not captcha:
            print("Captcha required")
            return

        result = scraper.search_by_case_details(
            state_code=args.state_code,
            dist_code=args.dist_code,
            court_complex_code=args.complex_code,
            est_code=args.est_code,
            case_type=args.case_type,
            case_number=args.case_number,
            year=args.year,
            captcha_code=captcha
        )
        display_case(result)

    # ================== cause lists ========================
    elif args.today or args.tomorrow or args.causelist:
        if not all([args.state_code, args.dist_code, args.complex_code, args.court_code]):
            print("Need --state-code, --dist-code, --complex-code, and --court-code")
            print(" Or use --interactive mode")
            return

        if args.today:
            date = datetime.now().strftime("%d-%m-%Y")
        elif args.tomorrow:
            date = (datetime.now() + timedelta(days=1)).strftime("%d-%m-%Y")
        else:
            date = args.date or datetime.now().strftime("%d-%m-%Y")

        # Retry logic for invalid captcha
        attempts = 2
        for i in range(attempts):
            captcha = args.captcha or get_captcha(scraper)
            if not captcha:
                print("Captcha required")
                return
            
            res = scraper.get_cause_list(
                state_code=args.state_code,
                dist_code=args.dist_code,
                court_complex_code=args.complex_code,
                court_code=args.court_code,
                date=date,
                captcha_code=captcha,
                case_type=args.kind,
                court_name_txt=location.get("court_name", "") if location else ""
            )
            
            if res.get("success") or "captcha" not in (res.get("message", "").lower()):
                result = res
                break
            
            print("Invalid captcha, try again…")
            args.captcha = None
        else:
            result = res

        display_cause_list(result)

        if args.stats:
            display_statistics(result)

        if args.excel:
            export_excel(result, args.excel)

        if args.download_pdf and result and result.get("success"):
            pdf_path = scraper.download_cause_list_pdf(
                state_code=args.state_code,
                dist_code=args.dist_code,
                court_complex_code=args.complex_code,
                court_code=args.court_code,
                date=date,
                out_dir="output",
                html=result.get("html"),
                case_type=args.kind,
                court_name_txt=location.get("court_name", "") if location else ""
            )
            if pdf_path:
                print(f"PDF downloaded: {pdf_path}")
            else:
                print("PDF not available")
    else:
        p.print_help()
        return

    # ================== saving =============================
    if result:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.output_json:
            save_json(result, args.output_json)
        else:
            save_json(result, f"result_{ts}.json")

        if args.output_text:
            save_text(result, args.output_text)

if __name__ == "__main__":
    main()